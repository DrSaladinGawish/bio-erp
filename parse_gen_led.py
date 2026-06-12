"""
parse_gen_led.py — Parse Gen_Led.xlsx 3 sheets, detect columns, output CSVs + analysis JSON

Sheets: Bnk_Led_251 (bank ledger), SAL_LED4 (sales ledger), PUR_LED3 (purchase ledger)

All use triple-journal GL pattern: each transaction spans 2-3 rows (Dr/Cr pairs).
"""

import json, csv, sys
from pathlib import Path
from collections import defaultdict

GEN_LED_PATH = Path(r"D:\flash memory\USB Drive\=====Bnk_Pur_Sal_Mod\Gen_Led.xlsx")
OUTPUT_DIR = Path(r"D:\Data_Sources\docs\gen_led_parsed")
DRY_RUN = "--dry-run" in sys.argv

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. pip install openpyxl")
    sys.exit(1)


SHEET_CONFIG = {
    "Bnk_Led_251": {
        "target_table": "bnk_transactions",
        "col_map": {
            "Index": "index",
            "Trnx_Typ": "transaction_type",
            "Transaction Date": "transaction_date",
            "Narration": "narration",
            "Transaction Reference": "transaction_reference",
            "Orig_Trx_Amt": "original_amount",
            "Cur_Nmn": "currency",
            "Account": "account_name",
            "Dr_Trx_Egp": "debit_egp",
            "Cr_Trx_Egp": "credit_egp",
        },
    },
    "SAL_LED4": {
        "target_table": "sales_invoices",
        "col_map": {
            "Trn_Type": "transaction_type",
            "invoice_number": "invoice_number",
            "issue_date": "issue_date",
            "client_name": "client_name",
            "description": "description",
            "unit_price": "unit_price",
            "quantity": "quantity",
            "electronic_code": "electronic_code",
            "cost_center": "cost_center",
            "account": "account_name",
            "DR_EGP": "debit_egp",
            "CR_EGP": "credit_egp",
        },
    },
    "PUR_LED3": {
        "target_table": "purchase_orders",
        "col_map": {
            "Trn_Type": "transaction_type",
            "Index": "index",
            "invoice_number": "invoice_number",
            "issue_date": "issue_date",
            "supplier_name": "supplier_name",
            "description": "description",
            "unit_price": "unit_price",
            "quantity": "quantity",
            "account": "account_name",
            "cost_center": "cost_center",
            "DR_EGP": "debit_egp",
            "CR_EGP": "credit_egp",
            "electronic_code": "electronic_code",
        },
    },
}


def find_header_row(ws, expected_keywords):
    """
    Scan first 10 rows for a header row that contains expected keywords.
    Returns (header_row_index, header_values) or (0, first_row).
    """
    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
        vals = [str(c).strip() if c else "" for c in row]
        joined = " ".join(vals)
        match_count = sum(1 for kw in expected_keywords if kw.lower() in joined.lower())
        if match_count >= 2:
            return i, list(row)
    return 0, list(next(ws.iter_rows(max_row=1, values_only=True)))


def normalize_headers(headers):
    """Map xlsx header names to canonical field names."""
    h_map = {}
    for raw in headers:
        s = str(raw).strip() if raw else ""
        key = s.lower().replace(".", "_").replace(" ", "_")
        key = key.replace("é", "e").replace("è", "e")
        # Explicit overrides
        overrides = {
            "trnx_typ_1": "trnx_typ",
            "orig_trx_amt": "orig_amt",
            "cur_nmn": "currency",
            "data_avg": "avg",
            "sup_mtbl6_name": "sub_ledger",
            "اسم الكود": "client_name",
            "الوصف": "description",
            "السعر": "unit_price",
            "الكمية": "quantity",
            "الرقم الإلكترونى": "electronic_code",
            "inv_master_مركز التكلفة": "cost_center",
            "sub_led_mtab1_name": "sub_ledger_name",
            "acccount": "account",
            "account": "account_name",
            "dr_egp": "debit_egp",
            "cr_egp": "credit_egp",
            "رقم الفاتورة": "invoice_number",
            "تاريخ الإصدار": "issue_date",
            "إسم الكود": "supplier_name",
        }
        if key in overrides:
            key = overrides[key]
        h_map[raw] = key
    return h_map


def extract_sheet(sheet_name, output_path):
    """Read sheet, find header, dump to CSV with normalized columns."""
    wb_src = openpyxl.load_workbook(GEN_LED_PATH, read_only=True, data_only=True)
    ws = wb_src[sheet_name]

    # Find header
    header_keywords = list(SHEET_CONFIG[sheet_name]["col_map"].keys())
    header_idx, raw_headers = find_header_row(ws, header_keywords)

    # Normalize headers
    h_map = normalize_headers(raw_headers)
    field_names = [h_map[h] for h in raw_headers]

    # Target mapping
    sheet_cfg = SHEET_CONFIG[sheet_name]
    col_config = sheet_cfg["col_map"]
    mapped_count = sum(
        1 for v in field_names if v in col_config.values() or v in col_config
    )

    print(f"  Header row: {header_idx + 1}")
    print(f"  Columns ({len(field_names)}): {field_names}")
    print(f"  Mapped to {sheet_cfg['target_table']}: {mapped_count}/{len(col_config)}")

    # Extract data rows
    data_rows = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if any(v for v in vals):
            data_rows.append({field_names[i]: vals[i] for i in range(len(vals))})

    # Write CSV
    if not DRY_RUN:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=field_names, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(data_rows)
        print(f"  CSV: {output_path} ({len(data_rows)} rows)")

    wb_src.close()
    return {
        "sheet": sheet_name,
        "target_table": sheet_cfg["target_table"],
        "total_rows": ws.max_row,
        "data_rows": len(data_rows),
        "columns": field_names,
        "mapped_columns": mapped_count,
        "expected_columns": len(col_config),
        "csv_path": str(output_path),
    }


def main():
    print("=== Gen_Led.xlsx Parser ===")
    if DRY_RUN:
        print("[DRY RUN — no files written]")

    if not GEN_LED_PATH.exists():
        print(f"ERROR: Gen_Led.xlsx not found at {GEN_LED_PATH}")
        print(
            "Found at USB root: D:\\flash memory\\USB Drive\\=====Bnk_Pur_Sal_Mod\\Gen_Led.xlsx"
        )
        sys.exit(1)

    wb = openpyxl.load_workbook(GEN_LED_PATH, read_only=True, data_only=True)
    print(f"\nSheets: {wb.sheetnames}")
    wb.close()

    results = []
    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_CONFIG:
            print(f"\nSkipping unknown sheet: {sheet_name}")
            continue
        print(f"\n--- {sheet_name} ---")
        csv_path = OUTPUT_DIR / f"{sheet_name}.csv"
        info = extract_sheet(sheet_name, csv_path)
        results.append(info)

    # Write analysis JSON
    analysis = {
        "source": str(GEN_LED_PATH),
        "total_sheets": len(results),
        "sheets": results,
        "summary": {
            r["sheet"]: {
                "rows": r["data_rows"],
                "target_table": r["target_table"],
                "columns": r["columns"],
                "mapping_accuracy": f"{r['mapped_columns']}/{r['expected_columns']}",
            }
            for r in results
        },
    }

    if not DRY_RUN:
        analysis_path = OUTPUT_DIR / "gen_led_analysis.json"
        with open(analysis_path, "w", encoding="utf-8") as f:
            json.dump(analysis, f, indent=2, ensure_ascii=False)
        print(f"\nAnalysis: {analysis_path}")

    print(f"\n{'=' * 50}")
    print(f"Total sheets parsed: {len(results)}")
    print(f"Output dir: {OUTPUT_DIR}")
    print(f"Next: Review gen_led_analysis.json, then build DB insert script")


if __name__ == "__main__":
    main()
